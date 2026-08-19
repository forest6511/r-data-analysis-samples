#!/usr/bin/env python3
"""第3章用サンプル sales.xlsx / sales.db の決定的生成（sales.csv から導出・乱数不使用）

- sales.xlsx
    - シート「売上データ」: sales.csv と同一の144行（日付は Excel の日付セル）
    - シート「店舗マスタ」: 店舗・地域・開店年（3行）
    - シート「月次報告」 : 表の上に説明行が3行ある報告書型シート（skip= の題材）
- sales.db (SQLite)
    - テーブル sales : sales.csv と同一の144行（日付は TEXT で保存される）
    - テーブル stores: 店舗マスタと同一
"""
import csv
import datetime as dt
import io
import re
import shutil
import sqlite3
import zipfile
from pathlib import Path

from openpyxl import Workbook

HERE = Path(__file__).parent
CH03 = HERE / "ch03"  # verify-code の検証フィクスチャ置き場（乖離防止のため同時更新する）

# xlsx は ZIP コンテナで、openpyxl は書き込み時刻を各エントリに埋め込む。
# バイト決定性のため、全エントリのタイムスタンプをこの固定値に正規化する
FIXED_ZIP_DATETIME = (2026, 1, 1, 0, 0, 0)


def save_xlsx_deterministic(wb: Workbook, path: Path) -> None:
    wb.properties.created = dt.datetime(*FIXED_ZIP_DATETIME)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fixed_iso = dt.datetime(*FIXED_ZIP_DATETIME).strftime("%Y-%m-%dT%H:%M:%SZ")
    with zipfile.ZipFile(buf) as src, zipfile.ZipFile(
        path, "w", compression=zipfile.ZIP_DEFLATED
    ) as dst:
        for name in sorted(src.namelist()):
            data = src.read(name)
            if name == "docProps/core.xml":
                # openpyxl は save 時に modified を現在時刻で上書きするため、ここで固定値に戻す
                data = re.sub(
                    rb"(<dcterms:modified[^>]*>)[^<]*(</dcterms:modified>)",
                    lambda m: m.group(1) + fixed_iso.encode() + m.group(2),
                    data,
                )
            info = zipfile.ZipInfo(name, date_time=FIXED_ZIP_DATETIME)
            info.compress_type = zipfile.ZIP_DEFLATED
            dst.writestr(info, data)

# --- 元データ読み込み ---
with open(HERE / "sales.csv", encoding="utf-8") as f:
    rows = list(csv.DictReader(f))

stores_master = [
    ("新宿店", "東京", 2008),
    ("大阪店", "大阪", 2012),
    ("オンライン", "全国", 2018),
]

# --- sales.xlsx ---
wb = Workbook()

ws = wb.active
ws.title = "売上データ"
ws.append(["日付", "店舗", "カテゴリ", "売上"])
for r in rows:
    d = dt.date.fromisoformat(r["日付"])
    ws.append([d, r["店舗"], r["カテゴリ"], int(r["売上"])])
for cell_row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
    cell_row[0].number_format = "yyyy-mm-dd"

ws2 = wb.create_sheet("店舗マスタ")
ws2.append(["店舗", "地域", "開店年"])
for row in stores_master:
    ws2.append(list(row))

ws3 = wb.create_sheet("月次報告")
ws3.append(["月次売上報告（2025年）"])
ws3.append(["作成: 営業企画部"])
ws3.append([])
ws3.append(["月", "売上合計"])
monthly = {}
for r in rows:
    m = int(r["日付"][5:7])
    monthly[m] = monthly.get(m, 0) + int(r["売上"])
for m in range(1, 13):
    ws3.append([m, monthly[m]])

save_xlsx_deterministic(wb, HERE / "sales.xlsx")

# --- sales.db ---
db_path = HERE / "sales.db"
db_path.unlink(missing_ok=True)
con = sqlite3.connect(db_path)
con.execute('CREATE TABLE sales ("日付" TEXT, "店舗" TEXT, "カテゴリ" TEXT, "売上" INTEGER)')
con.executemany(
    "INSERT INTO sales VALUES (?, ?, ?, ?)",
    [(r["日付"], r["店舗"], r["カテゴリ"], int(r["売上"])) for r in rows],
)
con.execute('CREATE TABLE stores ("店舗" TEXT, "地域" TEXT, "開店年" INTEGER)')
con.executemany("INSERT INTO stores VALUES (?, ?, ?)", stores_master)
con.commit()
con.close()

# --- 検証フィクスチャ（samples/ch03/）へ同期 ---
if CH03.is_dir():
    shutil.copy2(HERE / "sales.xlsx", CH03 / "sales.xlsx")
    shutil.copy2(HERE / "sales.db", CH03 / "sales.db")
    synced = " (ch03/ へ同期済み)"
else:
    synced = ""

print(f"{len(rows)} rows -> sales.xlsx / sales.db{synced}")
